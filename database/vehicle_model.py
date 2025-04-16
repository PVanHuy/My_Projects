from datetime import datetime
import os
import logging

# Cấu hình logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

class VehicleExporter:
    """Class chuyên về xuất dữ liệu xe"""
    
    @staticmethod
    def export_to_csv(vehicles, filename):
        """
        Xuất dữ liệu xe ra file CSV với định dạng đẹp
        
        Args:
            vehicles (list): Danh sách xe cần xuất
            filename (str): Đường dẫn file đích
            
        Returns:
            bool: True nếu xuất thành công, False nếu thất bại
        """
        try:
            import csv
            
            with open(filename, mode='w', newline='', encoding='utf-8-sig') as file:
                writer = csv.writer(file)
                
                # Viết header với định dạng tiếng Việt
                writer.writerow([
                    "STT", "Biển số", "Chủ xe", "Số điện thoại", 
                    "Loại xe", "Hãng xe", "Màu xe", "Thời gian đăng ký", "Ghi chú"
                ])
                
                # Viết dữ liệu
                for i, vehicle in enumerate(vehicles, start=1):
                    writer.writerow([
                        i,
                        vehicle.get("plate", ""),
                        vehicle.get("owner", ""),
                        vehicle.get("phone", ""),
                        vehicle.get("type", ""),
                        vehicle.get("brand", ""),
                        vehicle.get("color", "Không có thông tin"),
                        vehicle.get("timestamp", ""),
                        vehicle.get("notes", "")
                    ])
            
            logging.info(f"Exported data to CSV: {filename}")
            return True
            
        except Exception as e:
            logging.error(f"Error exporting to CSV: {str(e)}")
            return False

    @staticmethod
    def export_to_excel(vehicles, filename):
        """
        Xuất dữ liệu xe ra file Excel với định dạng đẹp và style
        
        Args:
            vehicles (list): Danh sách xe cần xuất
            filename (str): Đường dẫn file đích
            
        Returns:
            bool: True nếu xuất thành công, False nếu thất bại
        """
        try:
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils.dataframe import dataframe_to_rows
            from openpyxl.utils import get_column_letter
            
            # Tạo DataFrame từ data
            data = []
            for i, vehicle in enumerate(vehicles, start=1):
                data.append([
                    i,
                    vehicle.get("plate", ""),
                    vehicle.get("owner", ""),
                    vehicle.get("phone", ""),
                    vehicle.get("type", ""),
                    vehicle.get("brand", ""),
                    vehicle.get("color", "Không có thông tin"),
                    vehicle.get("timestamp", ""),
                    vehicle.get("notes", "")
                ])
                
            # Tạo DataFrame
            columns = ["STT", "Biển số", "Chủ xe", "Số điện thoại", 
                      "Loại xe", "Hãng xe", "Màu xe", "Thời gian đăng ký", "Ghi chú"]
            
            df = pd.DataFrame(data, columns=columns)
            
            # Cố gắng sử dụng openpyxl để tạo file Excel có định dạng đẹp
            try:
                # Tạo workbook và worksheet
                wb = Workbook()
                ws = wb.active
                ws.title = "Danh sách xe"
                
                # Thêm tiêu đề
                ws.merge_cells('A1:I1')
                title_cell = ws['A1']
                title_cell.value = "DANH SÁCH XE ĐÃ ĐĂNG KÝ"
                title_cell.font = Font(name='Arial', size=16, bold=True, color="0000FF")
                title_cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Thêm thông tin thời gian xuất
                ws.merge_cells('A2:I2')
                ws['A2'] = f"Thời gian xuất: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
                ws['A2'].font = Font(name='Arial', size=10, italic=True)
                ws['A2'].alignment = Alignment(horizontal='center')
                
                # Thêm dữ liệu từ DataFrame từ dòng 4
                for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 4):
                    for c_idx, value in enumerate(row, 1):
                        cell = ws.cell(row=r_idx, column=c_idx, value=value)
                        
                        # Định dạng header (dòng đầu tiên)
                        if r_idx == 4:
                            cell.font = Font(bold=True, color="FFFFFF")
                            cell.fill = PatternFill(start_color="2E5077", end_color="2E5077", fill_type="solid")
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                        else:
                            # Định dạng cho dữ liệu
                            cell.alignment = Alignment(vertical='center')
                            
                            # Đổi màu hàng chẵn lẻ để dễ đọc
                            if (r_idx - 4) % 2 == 1:  # Dòng lẻ (không tính header)
                                cell.fill = PatternFill(start_color="F6F4F0", end_color="F6F4F0", fill_type="solid")
                
                # Định dạng cột STT và biển số ra giữa
                for row in range(5, ws.max_row + 1):
                    ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
                    ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
                
                # Định dạng kích thước cột
                column_widths = {
                    'A': 5,   # STT
                    'B': 15,  # Biển số
                    'C': 25,  # Chủ xe
                    'D': 15,  # SĐT
                    'E': 15,  # Loại xe
                    'F': 15,  # Hãng xe
                    'G': 15,  # Màu xe
                    'H': 20,  # Thời gian
                    'I': 30,  # Ghi chú
                }
                
                for col, width in column_widths.items():
                    ws.column_dimensions[col].width = width
                
                # Thêm border cho tất cả các ô có dữ liệu
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                for row in range(4, ws.max_row + 1):
                    for col in range(1, 10):  # A-I
                        ws.cell(row=row, column=col).border = thin_border
                
                # Lưu workbook
                wb.save(filename)
                logging.info(f"Exported formatted Excel file: {filename}")
                return True
                
            except ImportError:
                # Fallback nếu không có openpyxl
                df.to_excel(filename, sheet_name="Danh sách xe", index=False)
                logging.info(f"Exported basic Excel file: {filename}")
                return True
                
        except Exception as e:
            logging.error(f"Error exporting to Excel: {str(e)}")
            return False

    @staticmethod
    def export_vehicle_info_to_pdf(vehicle, filename):
        """
        Xuất thông tin chi tiết một xe ra file PDF
        
        Args:
            vehicle (dict): Thông tin xe cần xuất
            filename (str): Đường dẫn file đích
            
        Returns:
            bool: True nếu xuất thành công, False nếu thất bại
        """
        try:
            from reportlab.pdfgen import canvas
            from reportlab.lib.pagesizes import A4
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            from reportlab.lib.units import cm
            import textwrap
            
            # Đăng ký font Unicode để hỗ trợ tiếng Việt
            try:
                # Thử tải font Arial Unicode
                font_path = "fonts/arial_unicode.ttf"
                
                # Kiểm tra nếu thư mục fonts chưa tồn tại thì tạo mới
                if not os.path.exists("fonts"):
                    os.makedirs("fonts")
                
                # Nếu font không có sẵn, sử dụng DejaVuSans mặc định
                if not os.path.exists(font_path):
                    logging.warning("Arial Unicode font not found, using DejaVu Sans")
                    font_path = "fonts/DejaVuSans.ttf"
                    
                    # Tải DejaVuSans từ reportlab nếu cần
                    from reportlab.pdfbase.ttfonts import TTFQuery
                    if not os.path.exists(font_path):
                        import reportlab
                        reportlab_path = os.path.dirname(reportlab.__file__)
                        dejavu_path = os.path.join(reportlab_path, 'fonts/DejaVuSans.ttf')
                        if os.path.exists(dejavu_path):
                            import shutil
                            shutil.copy(dejavu_path, font_path)
                
                pdfmetrics.registerFont(TTFont('UnicodeFont', font_path))
            except Exception as e:
                logging.warning(f"Could not register Unicode font: {e}")
            
            # Tạo canvas mới
            c = canvas.Canvas(filename, pagesize=A4)
            width, height = A4
            
            # Set font cho toàn bộ tài liệu
            c.setFont('UnicodeFont' if 'UnicodeFont' in pdfmetrics.getRegisteredFontNames() else 'Helvetica', 12)
            
            # Thêm tiêu đề
            c.setFont('UnicodeFont' if 'UnicodeFont' in pdfmetrics.getRegisteredFontNames() else 'Helvetica-Bold', 18)
            c.drawCentredString(width/2, height - 2*cm, "THÔNG TIN CHI TIẾT XE")
            
            # Vẽ đường kẻ dưới tiêu đề
            c.line(2*cm, height - 2.5*cm, width - 2*cm, height - 2.5*cm)
            
            # Đặt lại font
            c.setFont('UnicodeFont' if 'UnicodeFont' in pdfmetrics.getRegisteredFontNames() else 'Helvetica', 12)
            
            # Thông tin biển số (nổi bật)
            c.setFont('UnicodeFont' if 'UnicodeFont' in pdfmetrics.getRegisteredFontNames() else 'Helvetica-Bold', 14)
            c.drawString(2*cm, height - 4*cm, f"Biển số: {vehicle.get('plate', 'N/A')}")
            
            # Thông tin cơ bản
            c.setFont('UnicodeFont' if 'UnicodeFont' in pdfmetrics.getRegisteredFontNames() else 'Helvetica', 12)
            y_position = height - 5*cm
            line_height = 0.8*cm
            
            details = [
                f"Chủ xe: {vehicle.get('owner', 'N/A')}",
                f"Số điện thoại: {vehicle.get('phone', 'N/A')}",
                f"Loại xe: {vehicle.get('type', 'N/A')}",
                f"Hãng xe: {vehicle.get('brand', 'N/A')}",
                f"Màu xe: {vehicle.get('color', 'N/A')}",
                f"Thời gian đăng ký: {vehicle.get('timestamp', 'N/A')}",
            ]
            
            for line in details:
                c.drawString(2*cm, y_position, line)
                y_position -= line_height
            
            # Xử lý ghi chú nhiều dòng
            notes = vehicle.get('notes', 'Không có ghi chú')
            c.drawString(2*cm, y_position, "Ghi chú:")
            y_position -= line_height
            
            # Wrap text nếu dài
            for line in textwrap.wrap(notes, width=70):
                c.drawString(3*cm, y_position, line)
                y_position -= line_height
            
            # Thêm thời gian xuất báo cáo
            c.setFont('UnicodeFont' if 'UnicodeFont' in pdfmetrics.getRegisteredFontNames() else 'Helvetica-Oblique', 10)
            c.drawString(2*cm, 2*cm, f"Xuất báo cáo ngày: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
            
            # Lưu file
            c.save()
            
            logging.info(f"Exported vehicle info to PDF: {filename}")
            return True
            
        except Exception as e:
            logging.error(f"Error exporting to PDF: {str(e)}")
            return False
            
    @staticmethod
    def export_to_html(vehicles, filename):
        """
        Xuất danh sách xe ra file HTML
        
        Args:
            vehicles (list): Danh sách xe cần xuất
            filename (str): Đường dẫn file đích
            
        Returns:
            bool: True nếu xuất thành công, False nếu thất bại
        """
        try:
            html_template = """<!DOCTYPE html>
<html lang="vi">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Danh sách xe đăng ký</title>
    <style>
        body {
            font-family: Arial, sans-serif;
            margin: 20px;
            color: #212529;
        }
        h1 {
            color: #2E5077;
            text-align: center;
        }
        .export-time {
            text-align: center;
            font-style: italic;
            margin-bottom: 20px;
            color: #6c757d;
        }
        table {
            width: 100%;
            border-collapse: collapse;
            margin-top: 20px;
        }
        th {
            background-color: #2E5077;
            color: white;
            text-align: left;
            padding: 12px;
        }
        td {
            padding: 10px;
            border-bottom: 1px solid #ddd;
        }
        tr:nth-child(even) {
            background-color: #f6f4f0;
        }
        tr:hover {
            background-color: #e9ecef;
        }
        .center {
            text-align: center;
        }
    </style>
</head>
<body>
    <h1>DANH SÁCH XE ĐÃ ĐĂNG KÝ</h1>
    <div class="export-time">Thời gian xuất: EXPORT_TIME</div>
    
    <table>
        <thead>
            <tr>
                <th class="center">STT</th>
                <th class="center">Biển số</th>
                <th>Chủ xe</th>
                <th>Số điện thoại</th>
                <th>Loại xe</th>
                <th>Hãng xe</th>
                <th>Màu xe</th>
                <th>Thời gian đăng ký</th>
                <th>Ghi chú</th>
            </tr>
        </thead>
        <tbody>
            TABLE_ROWS
        </tbody>
    </table>
</body>
</html>
"""
            # Tạo nội dung table rows
            table_rows = ""
            for i, vehicle in enumerate(vehicles, start=1):
                table_rows += f"""
            <tr>
                <td class="center">{i}</td>
                <td class="center">{vehicle.get("plate", "")}</td>
                <td>{vehicle.get("owner", "")}</td>
                <td>{vehicle.get("phone", "")}</td>
                <td>{vehicle.get("type", "")}</td>
                <td>{vehicle.get("brand", "")}</td>
                <td>{vehicle.get("color", "Không có thông tin")}</td>
                <td>{vehicle.get("timestamp", "")}</td>
                <td>{vehicle.get("notes", "")}</td>
            </tr>"""
            
            # Thay thế placeholder
            export_time = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
            html_content = html_template.replace("EXPORT_TIME", export_time).replace("TABLE_ROWS", table_rows)
            
            # Lưu file HTML
            with open(filename, 'w', encoding='utf-8') as f:
                f.write(html_content)
                
            logging.info(f"Exported data to HTML: {filename}")
            return True
            
        except Exception as e:
            logging.error(f"Error exporting to HTML: {str(e)}")
            return False


def create_vehicle(plate, owner, phone, vehicle_type, brand, color="Không có thông tin", notes=""):
    """
    Creates a new vehicle entry with the given information.
    
    Args:
        plate (str): License plate number
        owner (str): Vehicle owner name
        phone (str): Owner's phone number
        vehicle_type (str): Type of vehicle (e.g., Sedan, SUV)
        brand (str): Vehicle manufacturer (e.g., Toyota, Honda)
        color (str, optional): Vehicle color. Defaults to "Không có thông tin".
        notes (str, optional): Additional notes. Defaults to "".
    
    Returns:
        dict: A new vehicle entry with all information and timestamp
    """
    # Sử dụng database manager để thêm xe
    from database.db_manager import DatabaseManager
    db = DatabaseManager()
    success, result = db.add_vehicle(
        plate=plate,
        owner=owner,
        phone=phone,
        vehicle_type=vehicle_type,
        brand=brand,
        color=color,
        notes=notes
    )
    
    if success:
        return result
    else:
        raise Exception(f"Không thể thêm xe: {result}")

def update_vehicle(vehicle_data, plate, **updates):
    """
    Updates an existing vehicle in the data list.
    
    Args:
        vehicle_data (list): List of vehicle dictionaries (không còn sử dụng)
        plate (str): License plate number to identify the vehicle
        **updates: Keyword arguments for fields to update
    
    Returns:
        bool: True if the vehicle was found and updated, False otherwise
    """
    # Sử dụng database manager để cập nhật xe
    from database.db_manager import DatabaseManager
    db = DatabaseManager()
    success, result = db.update_vehicle(plate, **updates)
    
    if success:
        # Cập nhật lại vehicle_data nếu cần
        if isinstance(vehicle_data, list) and isinstance(result, dict):
            for i, vehicle in enumerate(vehicle_data):
                if vehicle["plate"] == plate:
                    vehicle_data[i] = result
                    break
        return True
    else:
        return False

def delete_vehicle(vehicle_data, plate):
    """
    Deletes a vehicle from the data list.
    
    Args:
        vehicle_data (list): List of vehicle dictionaries
        plate (str): License plate number to identify the vehicle
    
    Returns:
        bool: True if the vehicle was found and deleted, False otherwise
    """
    # Sử dụng database manager để xóa xe
    from database.db_manager import DatabaseManager
    db = DatabaseManager()
    success, result = db.delete_vehicle(plate)
    
    if success:
        # Cập nhật lại vehicle_data nếu cần
        if isinstance(vehicle_data, list):
            for i, vehicle in enumerate(vehicle_data):
                if vehicle["plate"] == plate:
                    del vehicle_data[i]
                    break
        return True
    else:
        return False

def find_vehicle(vehicle_data, plate=None, owner=None, phone=None):
    """
    Finds vehicles matching the provided criteria.
    
    Args:
        vehicle_data (list): List of vehicle dictionaries
        plate (str, optional): License plate to search for. Defaults to None.
        owner (str, optional): Owner name to search for. Defaults to None.
        phone (str, optional): Phone number to search for. Defaults to None.
    
    Returns:
        list: List of vehicles matching the criteria
    """
    # Sử dụng database manager để tìm kiếm xe
    from database.db_manager import DatabaseManager
    db = DatabaseManager()
    search_text = None
    
    if plate:
        search_text = plate
    elif owner:
        search_text = owner
    elif phone:
        search_text = phone
    
    results = db.search_vehicles(search_text=search_text)
    return results

# Các hàm xuất dữ liệu đơn giản hơn để giữ khả năng tương thích ngược
def export_to_csv(vehicle_data, filename):
    return VehicleExporter.export_to_csv(vehicle_data, filename)

def export_to_excel(vehicle_data, filename):
    return VehicleExporter.export_to_excel(vehicle_data, filename)

def export_vehicle_info_to_pdf(vehicle, filename):
    return VehicleExporter.export_vehicle_info_to_pdf(vehicle, filename)

def export_to_html(vehicle_data, filename):
    return VehicleExporter.export_to_html(vehicle_data, filename)

def import_from_csv(filename):
    """
    Imports vehicle data from a CSV file.
    
    Args:
        filename (str): Path to the CSV file
    
    Returns:
        list: List of vehicle dictionaries imported from the CSV
    """
    try:
        import csv
        
        vehicle_data = []
        from database.db_manager import DatabaseManager
        db = DatabaseManager()
        
        with open(filename, mode='r', encoding='utf-8') as file:
            reader = csv.reader(file)
            next(reader)  # Skip header
            
            for row in reader:
                if len(row) >= 7:  # Ensure row has enough columns
                    plate = row[1]
                    owner = row[2]
                    phone = row[3]
                    vehicle_type = row[4]
                    brand = row[5]
                    color = row[6] if len(row) > 6 else "Không có thông tin"
                    notes = row[8] if len(row) > 8 else ""
                    
                    # Thêm vào database
                    success, result = db.add_vehicle(
                        plate=plate,
                        owner=owner,
                        phone=phone,
                        vehicle_type=vehicle_type,
                        brand=brand,
                        color=color,
                        notes=notes
                    )
                    
                    if success and isinstance(result, dict):
                        vehicle_data.append(result)
        
        return vehicle_data
    except Exception as e:
        logging.error(f"Error importing from CSV: {str(e)}")
        return []